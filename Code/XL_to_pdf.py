# Load Excel file
import time
import os.path
from datetime import datetime , date
import pandas as pd
from openpyxl.styles import Alignment , Font , Border , Side , Protection
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import global_var as GV


def Report_Generation():
    storage_path = '/media/chetan/14336adc-90be-4488-9a2d-9ed0b5e89990/Shift_Handover/'
    localtime = time.localtime(time.time())
    time_current = (str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(localtime[4]))
    today = date.today()
    date_now = today.strftime("%b-%d-%Y")
    file_path = storage_path + GV.Shift_incharge + '.xlsx'
    df = pd.DataFrame(
        columns=['Sr.No.' , 'Utility Equipment/Area Details' , 'Downtime in Numbers' , 'Downtime in Minutes' ,
                 'Downtime Form 17/16' , 'Remark'])
    if(os.path.isfile('/media/chetan/14336adc-90be-4488-9a2d-9ed0b5e89990/Shift_Handover/'+GV.Shift_incharge+'.xlsx')==True):
        writer = pd.ExcelWriter(file_path , mode='a',if_sheet_exists='new' , engine='openpyxl')
    else:
        writer= pd.ExcelWriter(file_path,mode='w',engine='openpyxl')
    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer , sheet_name=GV.date , index=False , startrow=4)

    writer.save()
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb[GV.date]

    thin_border = Border(left=Side(style='thin') ,
                         right=Side(style='thin') ,
                         top=Side(style='thin') ,
                         bottom=Side(style='thin'))
    logo = Image("/media/chetan/14336adc-90be-4488-9a2d-9ed0b5e89990/Shift_Handover/1.jpg")
    logo.height = 40
    logo.width = 80
    ws.row_dimensions[1].height = 40

    Alignment.vertical = 'top'
    ws['B5'].alignment = Alignment(wrap_text=True)
    ws['C5'].alignment = Alignment(wrap_text=True)
    ws['D5'].alignment = Alignment(wrap_text=True)
    ws['E5'].alignment = Alignment(wrap_text=True)
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.add_image(logo , "A1")
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.row_dimensions[5].height = 45
    ws.row_dimensions[3].height = 27

    ws.merge_cells('B1:F1')
    ws.merge_cells('B2:C2')
    ws.merge_cells('B3:C3')
    ws.merge_cells('D2:D3')
    ws.merge_cells('E2:E3')
    lst = ['HT Yard' , 'DG Yard' , 'Air Compressor Area' , 'Fire House Area' , 'PT L-I' , 'PH L-II' , 'STP Plant' ,
           'Softner Plant']
    for i in range(1 , 13):
        ws.cell(row=5 + i , column=1 , value=i)
    for j in range(len(lst)):
        ws.cell(row=6 + j , column=2 , value=lst[j])
    ws.cell(row=1 , column=2 , value="Utility Maintenance- Shift Handover Report ")
    cell = ws.cell(row=1 , column=2)
    cell.alignment = Alignment(horizontal='center' , vertical='center')
    cell.font = Font(bold=True)
    ws.cell(row=2 , column=1 , value="Date :")
    ws.cell(row=2 , column=2 , value=GV.date)
    ws.cell(row=3 , column=1 , value="Shift incharge :")
    ws.cell(row=3 , column=2 , value=GV.Shift_incharge)
    ws.cell(row=2 , column=4 , value="Shift :")
    ws.cell(row=2 , column=5 , value=GV.Shift)
    ws.cell(row=18 , column=1 , value="Shift major activity :-")
    ws.cell(row=19 , column=1 , value="Time")
    ws.cell(row=19 , column=2 , value="Plan")
    ws.cell(row=19 , column=4 , value="Action")
    ws.merge_cells('B19:C19')
    ws.merge_cells('D19:F19')
    ws.merge_cells('A18:F18')
    Time = ['07:00 AM' , '08:00 AM' , '09:00 AM' , '10:00 AM' , '11:00 AM' , '12:00 PM' , '01:00 PM' , '02:00 PM' ,
            '03:00 PM' , '04:00 PM' , '05:00 PM' , '06:00 PM' , '07:00 PM' , '08:00 PM' , '09:00 PM' , '10:00 PM' ,
            '11:00 PM' , '12:00 AM' , '01:00 AM' , '02:00 AM' , '03:00 AM' , '04:00 AM' , '05:00 AM' , '06:00 AM' ,
            '07:00 AM']
    for i in range(len(Time)):
        ws.cell(row=20 + i , column=1 , value=Time[i])
    for j in range(20 , 45):
        ws.merge_cells('B' + str(j) + ':' + 'C' + str(j))
        ws.merge_cells('D' + str(j) + ':' + 'F' + str(j))

    ws.cell(row=45 , column=1 , value="Handover By :")
    ws.cell(row=46 , column=1 , value="Name:" + GV.Handoverby + " \n Sign: ")

    ws.cell(row=45 , column=5 , value="Verifed By: ")
    ws.merge_cells('A45:B45')
    ws.merge_cells('A46:B47')
    ws.merge_cells('E45:F45')
    ws.cell(row=45 , column=3 , value="Takeover By :")
    ws.cell(row=46 , column=3 , value="Name:" + GV.Takeoverby + " \n Sign: ")
    ws.cell(row=46 , column=5 , value="Approved By:")
    ws.merge_cells('C45:D45')
    ws.merge_cells('C46:D47')
    ws.merge_cells('E46:F46')
    # ws.cell(row=46 , column=1).border = thin_border
    # ws.cell(row=1 , column=2).border = thin_border
    for a in range(1 , 7):
        for b in range(1 , 48):
            ws.cell(row=b , column=a).border = thin_border
    for i in range(len(GV.Tblcheck)):
        ws.cell(row=6 + i , column=3 , value=GV.Tblcheck[i][2])
        ws.cell(row=6 + i , column=4 , value=GV.Tblcheck[i][3])
        ws.cell(row=6 + i , column=5 , value=GV.Tblcheck[i][4])
        ws.cell(row=6 + i , column=6 , value=GV.Tblcheck[i][5])
    for j in range(len(GV.readcheck)):
        ws.cell(row=20 + j , column=2 , value=GV.readcheck[j][1])
        ws.cell(row=20 + j , column=4 , value=GV.readcheck[j][2])

    GV.Tblcheck = []
    GV.readcheck = []
    wb.save(file_path)

    # os.system('unoconv -f pdf ' + file_path + '')
    # print('Done' , file_path)

# Report_Generation()
